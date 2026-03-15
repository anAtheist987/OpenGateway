# Copyright 2026 Tsinghua University
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# This file was created by Tsinghua University and is not part of
# the original agentgateway project by Solo.io.

"""Shared MCP server for procurement/foreign/safety document extraction."""
import base64
import html as html_lib
import io
import json
import os
import re
from typing import Any

from dotenv import load_dotenv
from litellm import acompletion
from mcp.server.fastmcp import FastMCP

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
ROOT_ENV_PATH = os.path.join(PROJECT_ROOT, '.env')
LOCAL_ENV_PATH = os.path.join(CURRENT_DIR, '.env')

if os.path.exists(ROOT_ENV_PATH):
    load_dotenv(ROOT_ENV_PATH)
if os.path.exists(LOCAL_ENV_PATH):
    load_dotenv(LOCAL_ENV_PATH, override=True)
if not os.path.exists(ROOT_ENV_PATH) and not os.path.exists(LOCAL_ENV_PATH):
    load_dotenv()

mcp = FastMCP('dept_doc_reader')
DEFAULT_PORTAL_DATA_PATH = '/mnt/ssd2/cyh/Agentgateway-thu/reimbursement_portal/reimbursement_mock_data.json'
SUPPORTED_DEPARTMENTS = {
    'procurement': '采购与集采管理中心',
    'foreign': '外事与出入境管理办公室',
    'safety': '安全与海外风险管理中心',
}
BLOCKED_DEPARTMENTS = {'finance', 'infosec'}


def _normalize_api_base(url: str | None) -> str | None:
    if not url:
        return None
    clean = url.strip().rstrip('/')
    if not clean:
        return None
    return clean


def _resolve_document_llm_config() -> dict[str, str]:
    mode = os.getenv('DOCUMENT_LLM_MODE', 'api').strip().lower()
    if mode == 'local':
        model = os.getenv(
            'DOCUMENT_LOCAL_MODEL',
            os.getenv('LITELLM_MODEL', 'openai/qwen3.5-27b-fp8'),
        ).strip()
        api_base = _normalize_api_base(
            os.getenv('DOCUMENT_LOCAL_API_BASE', 'http://localhost:8888')
        )
        api_key = os.getenv('DOCUMENT_LOCAL_API_KEY', 'EMPTY').strip()
    else:
        model = os.getenv(
            'DOCUMENT_API_MODEL',
            os.getenv('LITELLM_MODEL', 'openai/qwen3.5-27b-fp8'),
        ).strip()
        api_base = _normalize_api_base(
            os.getenv('DOCUMENT_API_BASE', 'https://api.anthropic.com')
        )
        api_key = os.getenv('DOCUMENT_API_KEY', '').strip()

    if api_base and 'claude' in model.lower() and api_base.endswith('/v1'):
        api_base = api_base[:-3]
    return {
        'mode': mode,
        'model': model,
        'api_base': api_base or '',
        'api_key': api_key,
    }


def _pdf_to_text(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return '\n'.join(page.extract_text() or '' for page in reader.pages).strip()


def _docx_to_text(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return '\n'.join(p.text for p in doc.paragraphs).strip()


def _xlsx_to_text(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            line = '\t'.join(str(c) if c is not None else '' for c in row).strip()
            if line:
                parts.append(line)
    wb.close()
    return '\n'.join(parts).strip()


def _file_to_text(data: bytes, extension: str) -> str:
    ext = extension.lower().lstrip('.')
    if ext == 'pdf':
        return _pdf_to_text(data)
    if ext in ('docx',):
        return _docx_to_text(data)
    if ext == 'xlsx':
        return _xlsx_to_text(data)
    raise ValueError('Unsupported format: .%s. Supported: .pdf, .docx, .xlsx' % ext)


def _html_to_text(html_content: str) -> str:
    text = re.sub(r'<script[\s\S]*?</script>', ' ', html_content, flags=re.I)
    text = re.sub(r'<style[\s\S]*?</style>', ' ', text, flags=re.I)
    text = re.sub(r'<[^>]+>', '\n', text)
    text = html_lib.unescape(text)
    text = re.sub(r'\n{2,}', '\n', text)
    return text.strip()


def _portal_data_path() -> str:
    return os.getenv('REIMBURSEMENT_MOCK_DATA_PATH', DEFAULT_PORTAL_DATA_PATH)


def _load_portal_data() -> dict[str, Any]:
    path = _portal_data_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f'Portal data file not found: {path}')
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _validate_department_id(department_id: str) -> str | None:
    dep = (department_id or '').strip().lower()
    if dep in BLOCKED_DEPARTMENTS:
        return (
            f'department_id={dep} must use dedicated agent. '
            'Call finance_document_agent or infosec_document_agent instead.'
        )
    if dep not in SUPPORTED_DEPARTMENTS:
        return (
            f'Unsupported department_id={dep}. '
            f'Supported: {", ".join(sorted(SUPPORTED_DEPARTMENTS))}'
        )
    return None


def _filter_notices(data: dict[str, Any], department_id: str | None = None) -> list[dict[str, Any]]:
    notices = data.get('notices', [])
    if not department_id:
        return [
            n for n in notices if str(n.get('department_id', '')).lower() in SUPPORTED_DEPARTMENTS
        ]
    dep = department_id.strip().lower()
    return [n for n in notices if str(n.get('department_id', '')).lower() == dep]


def _compose_notice_text(notice: dict[str, Any], include_attachment: bool) -> str:
    parts = [
        f"NoticeID: {notice.get('id', '')}",
        f"Title: {notice.get('title', '')}",
        f"Department: {notice.get('department_name', '')}",
        f"DepartmentID: {notice.get('department_id', '')}",
        f"Category: {notice.get('category', '')}",
        f"PublishDate: {notice.get('publish_date', '')}",
        f"Body: {notice.get('content_text', '')}",
    ]
    if include_attachment and notice.get('attachment_name'):
        parts.append(f"Attachment: {notice.get('attachment_name')}")
        attachment_lines = notice.get('attachment_lines') or []
        if attachment_lines:
            parts.append('Attachment summary:')
            parts.extend(str(x) for x in attachment_lines)
    return '\n'.join(parts).strip()


EXTRACT_SCHEMA = """
返回一个 JSON 对象，且仅包含这些键：
- constraints
- recommendations
- process_steps
- material_checklist
- warnings
不要使用 Markdown 代码块包裹。
除通用英文缩写（如 OA、SLA）外，文本内容默认使用中文。
"""


def _strip_code_fence(text: str) -> str:
    text = (text or '').strip()
    if text.startswith('```'):
        lines = text.split('\n')
        if lines and lines[0].startswith('```'):
            lines = lines[1:]
        if lines and lines[-1].strip() == '```':
            lines = lines[:-1]
        text = '\n'.join(lines).strip()
    return text


def _parse_llm_json(text: str) -> dict[str, Any]:
    clean = _strip_code_fence(text)
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        start = clean.find('{')
        end = clean.rfind('}')
        if start != -1 and end != -1 and end > start:
            return json.loads(clean[start : end + 1])
        raise


async def _extract_with_llm(doc_text: str, categories: list[str] | None) -> dict[str, Any]:
    cfg = _resolve_document_llm_config()
    scope = ''
    if categories:
        scope = f"Only extract categories: {', '.join(categories)}."
    prompt = f"""Extract structured information from the department notice document below.
{scope}

Focus on: approval/filing/reimbursement process steps and material checklists.

Output requirements:
1) Respond in the same language as the document content;
2) Keep necessary abbreviations (e.g. OA, SLA) as-is;
3) Output ONLY a JSON object matching the schema — no explanatory prose.

Document content:
---
{doc_text[:12000]}
---

{EXTRACT_SCHEMA}
"""
    try:
        kwargs: dict[str, Any] = {
            'model': cfg['model'],
            'messages': [
                {
                    'role': 'system',
                    'content': 'You are a corporate policy extraction assistant. Respond in the same language as the document content.',
                },
                {'role': 'user', 'content': prompt},
            ],
        }
        if cfg.get('api_base'):
            kwargs['api_base'] = cfg['api_base']
        if cfg.get('api_key'):
            kwargs['api_key'] = cfg['api_key']
        response = await acompletion(**kwargs)
        text = response.choices[0].message.content if response.choices else ''
        parsed = _parse_llm_json(text)
        parsed.setdefault('warnings', [])
        return parsed
    except Exception as e:
        return {
            'constraints': [],
            'recommendations': [],
            'process_steps': [],
            'material_checklist': [],
            'warnings': [f'Extraction failed: {e!s}'],
        }


def _empty_with_warning(msg: str) -> str:
    return json.dumps(
        {
            'constraints': [],
            'recommendations': [],
            'process_steps': [],
            'material_checklist': [],
            'warnings': [msg],
        },
        ensure_ascii=False,
        indent=2,
    )


@mcp.tool()
async def list_supported_departments() -> str:
    """List departments served by this shared reader."""
    return json.dumps(
        {
            'supported_department_ids': sorted(SUPPORTED_DEPARTMENTS.keys()),
            'blocked_department_ids': sorted(BLOCKED_DEPARTMENTS),
        },
        ensure_ascii=False,
        indent=2,
    )


@mcp.tool()
async def list_department_notices(
    department_id: str | None = None,
    limit: int = 20,
) -> str:
    """List notices for shared departments. department_id optional."""
    if department_id:
        error = _validate_department_id(department_id)
        if error:
            return json.dumps({'notices': [], 'warnings': [error]}, ensure_ascii=False, indent=2)
    try:
        data = _load_portal_data()
    except Exception as e:
        return json.dumps({'notices': [], 'warnings': [str(e)]}, ensure_ascii=False, indent=2)
    notices = _filter_notices(data, department_id=department_id)
    brief = [
        {
            'id': n.get('id'),
            'title': n.get('title'),
            'department_id': n.get('department_id'),
            'department_name': n.get('department_name'),
            'category': n.get('category'),
            'publish_date': n.get('publish_date'),
            'content_type': n.get('content_type'),
        }
        for n in notices
    ]
    brief = sorted(brief, key=lambda x: x.get('publish_date', ''), reverse=True)
    return json.dumps(
        {'total': len(brief), 'notices': brief[: max(1, min(limit, 100))]},
        ensure_ascii=False,
        indent=2,
    )


@mcp.tool()
async def search_department_notices(
    query: str,
    department_id: str,
    limit: int = 5,
) -> str:
    """Search notices by keyword within a specific supported department."""
    error = _validate_department_id(department_id)
    if error:
        return json.dumps({'notices': [], 'warnings': [error]}, ensure_ascii=False, indent=2)
    try:
        data = _load_portal_data()
    except Exception as e:
        return json.dumps({'notices': [], 'warnings': [str(e)]}, ensure_ascii=False, indent=2)

    notices = _filter_notices(data, department_id=department_id)
    q = (query or '').strip().lower()
    keywords = [k for k in re.split(r'[\s,，;；|/]+', q) if k]
    ranked: list[tuple[int, dict[str, Any], int]] = []
    for n in notices:
        text = ' '.join(
            [
                str(n.get('title', '')),
                str(n.get('content_text', '')),
                str(n.get('department_name', '')),
                str(n.get('category', '')),
                ' '.join(str(x) for x in (n.get('job_titles') or [])),
            ]
        ).lower()
        score = 0
        matched_keywords = 0
        if keywords:
            matched_keywords = sum(1 for kw in keywords if kw in text)
            if matched_keywords == 0:
                continue
            score += matched_keywords * 2
            if matched_keywords == len(keywords):
                score += 2
        ranked.append((score, n, matched_keywords))
    ranked.sort(key=lambda x: (x[0], x[1].get('publish_date', '')), reverse=True)
    result = [
        {
            'id': n.get('id'),
            'title': n.get('title'),
            'department_id': n.get('department_id'),
            'department_name': n.get('department_name'),
            'category': n.get('category'),
            'publish_date': n.get('publish_date'),
            'content_type': n.get('content_type'),
            'score': score,
            'matched_keywords': matched_keywords,
        }
        for score, n, matched_keywords in ranked[: max(1, min(limit, 20))]
    ]
    return json.dumps({'total': len(result), 'notices': result}, ensure_ascii=False, indent=2)


@mcp.tool()
async def extract_department_notices(
    department_id: str,
    notice_ids: list[str],
    categories: list[str] | None = None,
    include_attachment: bool = True,
) -> str:
    """Extract policy details from selected notice IDs in one supported department."""
    error = _validate_department_id(department_id)
    if error:
        return _empty_with_warning(error)
    if not notice_ids:
        return _empty_with_warning('notice_ids is empty')
    try:
        data = _load_portal_data()
    except Exception as e:
        return _empty_with_warning(str(e))
    notice_map = {
        str(n.get('id')): n for n in _filter_notices(data, department_id=department_id)
    }
    missing = []
    selected_texts = []
    for nid in notice_ids:
        item = notice_map.get(str(nid))
        if not item:
            missing.append(str(nid))
            continue
        selected_texts.append(_compose_notice_text(item, include_attachment))
    if not selected_texts:
        return _empty_with_warning(f'No valid notices found: {", ".join(missing)}')
    merged_text = '\n\n-----\n\n'.join(selected_texts)
    result = await _extract_with_llm(merged_text, categories)
    if missing:
        result.setdefault('warnings', [])
        result['warnings'].append(f'Missing IDs: {", ".join(missing)}')
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def extract_constraints_and_process(
    doc_text: str,
    categories: list[str] | None = None,
) -> str:
    """Extract structured policy information from plain text."""
    if not (doc_text and doc_text.strip()):
        return _empty_with_warning('Document text is empty')
    result = await _extract_with_llm(doc_text.strip(), categories)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def extract_from_file(
    file_content_base64: str,
    file_extension: str,
    categories: list[str] | None = None,
) -> str:
    """Extract structured policy information from file."""
    try:
        data = base64.b64decode(file_content_base64, validate=True)
    except Exception as e:
        return _empty_with_warning(f'Base64 decode failed: {e!s}')
    try:
        doc_text = _file_to_text(data, file_extension)
    except Exception as e:
        return _empty_with_warning(f'File parse failed: {e!s}')
    if not doc_text:
        return _empty_with_warning('No text extracted from file')
    result = await _extract_with_llm(doc_text, categories)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def extract_from_html(
    html_content: str,
    categories: list[str] | None = None,
) -> str:
    """Extract structured policy information from HTML."""
    if not (html_content and html_content.strip()):
        return _empty_with_warning('HTML content is empty')
    doc_text = _html_to_text(html_content)
    if not doc_text:
        return _empty_with_warning('No text after HTML cleanup')
    result = await _extract_with_llm(doc_text, categories)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def list_extraction_categories() -> str:
    """List extraction categories and process stages."""
    return json.dumps(
        {
            'categories': [
                'transport',
                'accommodation',
                'allowance',
                'approval',
                'audit',
                'reimbursement',
                'other',
            ],
            'stages': ['approval', 'filing', 'reimbursement'],
        },
        ensure_ascii=False,
        indent=2,
    )


if __name__ == '__main__':
    mcp.run(transport='stdio')

