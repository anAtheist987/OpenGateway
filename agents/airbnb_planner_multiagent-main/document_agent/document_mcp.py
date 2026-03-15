# Copyright 2026 Tsinghua University
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# This file was created by Tsinghua University and is not part of
# the original agentgateway project by Solo.io.

"""MCP server for document extraction: constraints, recommendations, process steps.

Extracts from travel policy / department documents into structured JSON
for use by the Host agent (constraints, recommendations, process_steps, material_checklist).
Supports plain text, PDF, Word (.docx), and Excel (.xlsx) as input.
"""
import base64
import html as html_lib
import io
import json
import os
import re
from typing import Any

from dotenv import load_dotenv
from litellm import acompletion
from mcp.server.fastmcp import FastMCP

# Env loading strategy for document agent:
# 1) Load project root .env (shared defaults)
# 2) Load document_agent/.env with override=True (document-agent specific overrides)
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
ROOT_ENV_PATH = os.path.join(PROJECT_ROOT, '.env')
LOCAL_ENV_PATH = os.path.join(CURRENT_DIR, '.env')

if os.path.exists(ROOT_ENV_PATH):
    load_dotenv(ROOT_ENV_PATH)
if os.path.exists(LOCAL_ENV_PATH):
    load_dotenv(LOCAL_ENV_PATH, override=True)
if not os.path.exists(ROOT_ENV_PATH) and not os.path.exists(LOCAL_ENV_PATH):
    load_dotenv()

mcp = FastMCP('document_extract')
DEFAULT_PORTAL_DATA_PATH = '/root/reimbursement_portal/reimbursement_mock_data.json'


def _normalize_api_base(url: str | None) -> str | None:
    if not url:
        return None
    clean = url.strip().rstrip('/')
    if not clean:
        return None
    return clean


def _resolve_document_llm_config() -> dict[str, str]:
    """Resolve document extraction LLM config from env.

    DOCUMENT_LLM_MODE controls provider selection:
    - local: use DOCUMENT_LOCAL_* vars
    - api: use DOCUMENT_API_* vars
    """
    mode = os.getenv('DOCUMENT_LLM_MODE', 'api').strip().lower()

    if mode == 'local':
        model = os.getenv(
            'DOCUMENT_LOCAL_MODEL',
            os.getenv('LITELLM_MODEL', 'openai/qwen3.5-27b-fp8'),
        ).strip()
        api_base = _normalize_api_base(
            os.getenv('DOCUMENT_LOCAL_API_BASE', 'http://localhost:8888')
        )
        api_key = os.getenv('DOCUMENT_LOCAL_API_KEY', 'EMPTY').strip()
    else:
        model = os.getenv(
            'DOCUMENT_API_MODEL',
            os.getenv('LITELLM_MODEL', 'openai/qwen3.5-27b-fp8'),
        ).strip()
        api_base = _normalize_api_base(
            os.getenv('DOCUMENT_API_BASE', 'https://api.anthropic.com')
        )
        api_key = os.getenv('DOCUMENT_API_KEY', '').strip()

    # Anthropic/Claude endpoints are usually rooted at host (without trailing /v1),
    # while the SDK/provider appends /v1/messages internally.
    if api_base and 'claude' in model.lower() and api_base.endswith('/v1'):
        api_base = api_base[:-3]

    return {
        'mode': mode,
        'model': model,
        'api_base': api_base or '',
        'api_key': api_key,
    }


def _pdf_to_text(data: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    return '\n'.join(
        page.extract_text() or ''
        for page in reader.pages
    ).strip()


def _docx_to_text(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return '\n'.join(p.text for p in doc.paragraphs).strip()


def _xlsx_to_text(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            line = '\t'.join(str(c) if c is not None else '' for c in row).strip()
            if line:
                parts.append(line)
    wb.close()
    return '\n'.join(parts).strip()


def _file_to_text(data: bytes, extension: str) -> str:
    ext = extension.lower().lstrip('.')
    if ext == 'pdf':
        return _pdf_to_text(data)
    if ext in ('docx',):
        return _docx_to_text(data)
    if ext == 'xlsx':
        return _xlsx_to_text(data)
    raise ValueError(
        '不支持的文件格式: .%s。支持: .pdf, .docx, .xlsx' % ext
    )


def _html_to_text(html_content: str) -> str:
    """Convert raw HTML into plain text."""
    text = re.sub(r'<script[\s\S]*?</script>', ' ', html_content, flags=re.I)
    text = re.sub(r'<style[\s\S]*?</style>', ' ', text, flags=re.I)
    text = re.sub(r'<[^>]+>', '\n', text)
    text = html_lib.unescape(text)
    text = re.sub(r'\n{2,}', '\n', text)
    return text.strip()


def _portal_data_path() -> str:
    return os.getenv('REIMBURSEMENT_MOCK_DATA_PATH', DEFAULT_PORTAL_DATA_PATH)


def _load_portal_data() -> dict[str, Any]:
    path = _portal_data_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f'公告数据文件不存在: {path}')
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _compose_notice_text(notice: dict[str, Any], include_attachment: bool) -> str:
    parts = [
        f"公告ID: {notice.get('id', '')}",
        f"标题: {notice.get('title', '')}",
        f"部门: {notice.get('department_name', '')}",
        f"类别: {notice.get('category', '')}",
        f"发布日期: {notice.get('publish_date', '')}",
        f"正文: {notice.get('content_text', '')}",
    ]
    if include_attachment and notice.get('attachment_name'):
        parts.append(f"附件名称: {notice.get('attachment_name')}")
        attachment_lines = notice.get('attachment_lines') or []
        if attachment_lines:
            parts.append('附件摘要:')
            parts.extend(str(x) for x in attachment_lines)
    return '\n'.join(parts).strip()

EXTRACT_SCHEMA = """
Return a single JSON object with exactly these keys (use empty arrays [] where nothing found):
- constraints: array of { "id": "c1", "category": "transport|accommodation|allowance|approval|audit|other", "rule_text": "硬性规则描述", "source": "文档来源", "source_excerpt": "原文片段" }
- recommendations: array of { "id": "r1", "category": "transport|accommodation|...", "condition_text": "软性推荐描述", "source": "文档来源", "source_excerpt": "原文片段" }
- process_steps: array of { "step_order": 1, "stage": "approval|filing|reimbursement", "name": "步骤名称", "materials": ["材料1"], "source": "文档来源" }
- material_checklist: array of strings, all materials needed (e.g. 出差申请单, 会议邀请函, 登机牌, 酒店水单, 发票)
- warnings: array of strings for missing/ambiguous clauses (e.g. "未找到「发票开具主体」相关条款")
Use UTF-8. No markdown code fence around JSON.
"""


def _strip_code_fence(text: str) -> str:
    text = (text or '').strip()
    if text.startswith('```'):
        lines = text.split('\n')
        if lines and lines[0].startswith('```'):
            lines = lines[1:]
        if lines and lines[-1].strip() == '```':
            lines = lines[:-1]
        text = '\n'.join(lines).strip()
    return text


def _parse_llm_json(text: str) -> dict[str, Any]:
    clean = _strip_code_fence(text)
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        # Try extracting the outermost JSON object from mixed text.
        start = clean.find('{')
        end = clean.rfind('}')
        if start != -1 and end != -1 and end > start:
            return json.loads(clean[start : end + 1])
        raise


async def _extract_with_llm(doc_text: str, categories: list[str] | None) -> dict[str, Any]:
    cfg = _resolve_document_llm_config()
    model = cfg['model']
    scope = ''
    if categories:
        scope = f" Only extract items in these categories: {', '.join(categories)}."
    prompt = f"""从以下企业差旅/制度文档中抽取结构化信息。

{scope}

文档内容：
---
{doc_text[:12000]}
---

{EXTRACT_SCHEMA}"""

    try:
        kwargs: dict[str, Any] = {
            'model': model,
            'messages': [{'role': 'user', 'content': prompt}],
        }
        if cfg.get('api_base'):
            kwargs['api_base'] = cfg['api_base']
        if cfg.get('api_key'):
            kwargs['api_key'] = cfg['api_key']

        response = await acompletion(
            **kwargs,
        )
        text = response.choices[0].message.content if response.choices else ''
        try:
            return _parse_llm_json(text)
        except json.JSONDecodeError:
            # One retry with stronger formatting instruction.
            retry_prompt = f"""你上一次输出不是合法 JSON。请严格按要求重写。

只允许输出一个 JSON 对象，禁止任何解释文字、前后缀、Markdown 代码块。
键必须且仅包含：constraints, recommendations, process_steps, material_checklist, warnings。

{EXTRACT_SCHEMA}

文档内容：
---
{doc_text[:12000]}
---
"""
            retry_kwargs = dict(kwargs)
            retry_kwargs['messages'] = [{'role': 'user', 'content': retry_prompt}]
            retry_resp = await acompletion(**retry_kwargs)
            retry_text = (
                retry_resp.choices[0].message.content if retry_resp.choices else ''
            )
            return _parse_llm_json(retry_text)
    except json.JSONDecodeError as e:
        return {
            'constraints': [],
            'recommendations': [],
            'process_steps': [],
            'material_checklist': [],
            'warnings': [f'LLM 返回非合法 JSON，解析失败: {e!s}'],
        }
    except Exception as e:
        return {
            'constraints': [],
            'recommendations': [],
            'process_steps': [],
            'material_checklist': [],
            'warnings': [f'抽取过程出错: {e!s}'],
        }


@mcp.tool()
async def extract_constraints_and_process(
    doc_text: str,
    categories: list[str] | None = None,
) -> str:
    """从差旅/制度文档正文中抽取：硬约束、软推荐、审批/报销流程步骤与材料清单。

    Args:
        doc_text: 文档正文（如差旅制度、审批流程说明的文本）。
        categories: 可选。限定抽取维度，如 ["transport", "accommodation", "approval", "reimbursement"]；
            不传则抽取全部维度。

    Returns:
        结构化 JSON 字符串，包含 constraints, recommendations, process_steps, material_checklist, warnings。
        每条均带 source/source_excerpt 便于可审计。
    """
    if not (doc_text and doc_text.strip()):
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': ['文档内容为空'],
            },
            ensure_ascii=False,
            indent=2,
        )
    result = await _extract_with_llm(doc_text.strip(), categories)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def extract_from_file(
    file_content_base64: str,
    file_extension: str,
    categories: list[str] | None = None,
) -> str:
    """从 PDF / Word(.docx) / Excel(.xlsx) 文件中抽取：硬约束、软推荐、审批/报销流程步骤与材料清单。

    先将文件转为文本再抽取，输出格式与 extract_constraints_and_process 一致。

    Args:
        file_content_base64: 文件内容的 Base64 编码字符串。
        file_extension: 文件扩展名，如 "pdf", "docx", "xlsx"（可带点，如 ".pdf"）。
        categories: 可选。限定抽取维度，如 ["transport", "accommodation", "approval", "reimbursement"]。

    Returns:
        结构化 JSON 字符串，包含 constraints, recommendations, process_steps, material_checklist, warnings。
    """
    try:
        data = base64.b64decode(file_content_base64, validate=True)
    except Exception as e:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': [f'Base64 解码失败: {e!s}'],
            },
            ensure_ascii=False,
            indent=2,
        )
    try:
        doc_text = _file_to_text(data, file_extension)
    except ValueError as e:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': [str(e)],
            },
            ensure_ascii=False,
            indent=2,
        )
    except Exception as e:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': [f'文件解析失败: {e!s}'],
            },
            ensure_ascii=False,
            indent=2,
        )
    if not doc_text:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': ['文件内容为空或无法提取文本'],
            },
            ensure_ascii=False,
            indent=2,
        )
    result = await _extract_with_llm(doc_text, categories)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def extract_from_html(
    html_content: str,
    categories: list[str] | None = None,
) -> str:
    """从 HTML 字符串中抽取：硬约束、软推荐、审批/报销流程步骤与材料清单。

    Args:
        html_content: 页面原始 HTML 内容（公告详情页或片段）。
        categories: 可选。限定抽取维度。

    Returns:
        结构化 JSON 字符串，包含 constraints, recommendations, process_steps, material_checklist, warnings。
    """
    if not (html_content and html_content.strip()):
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': ['HTML 内容为空'],
            },
            ensure_ascii=False,
            indent=2,
        )
    doc_text = _html_to_text(html_content)
    if not doc_text:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': ['HTML 清洗后无有效文本'],
            },
            ensure_ascii=False,
            indent=2,
        )
    result = await _extract_with_llm(doc_text, categories)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def list_portal_notices(
    department: str | None = None,
    category: str | None = None,
    limit: int = 20,
) -> str:
    """列出公告站 mock 数据中的公告摘要，供上游先筛选再抽取。

    Args:
        department: 可选，按部门名过滤（模糊匹配）。
        category: 可选，按类别过滤（模糊匹配）。
        limit: 返回条数上限，默认 20。
    """
    try:
        data = _load_portal_data()
    except Exception as e:
        return json.dumps({'notices': [], 'warnings': [str(e)]}, ensure_ascii=False, indent=2)

    notices = data.get('notices', [])
    dep_kw = (department or '').strip().lower()
    cat_kw = (category or '').strip().lower()

    filtered = []
    for n in notices:
        dep = str(n.get('department_name', '')).lower()
        cat = str(n.get('category', '')).lower()
        if dep_kw and dep_kw not in dep:
            continue
        if cat_kw and cat_kw not in cat:
            continue
        filtered.append(
            {
                'id': n.get('id'),
                'title': n.get('title'),
                'department_name': n.get('department_name'),
                'category': n.get('category'),
                'publish_date': n.get('publish_date'),
                'content_type': n.get('content_type'),
            }
        )

    filtered = sorted(filtered, key=lambda x: x.get('publish_date', ''), reverse=True)
    return json.dumps(
        {'total': len(filtered), 'notices': filtered[: max(1, min(limit, 100))]},
        ensure_ascii=False,
        indent=2,
    )


@mcp.tool()
async def search_portal_notices(
    query: str,
    department: str | None = None,
    category: str | None = None,
    limit: int = 5,
) -> str:
    """按关键词搜索公告站 mock 数据，返回候选公告列表（先检索、后抽取）。"""
    try:
        data = _load_portal_data()
    except Exception as e:
        return json.dumps({'notices': [], 'warnings': [str(e)]}, ensure_ascii=False, indent=2)

    notices = data.get('notices', [])
    q = (query or '').strip().lower()
    keywords = [k for k in re.split(r'[\s,，;；|/]+', q) if k]
    dep_kw = (department or '').strip().lower()
    cat_kw = (category or '').strip().lower()

    ranked = []
    for n in notices:
        text = ' '.join(
            [
                str(n.get('title', '')),
                str(n.get('content_text', '')),
                str(n.get('department_name', '')),
                str(n.get('category', '')),
                ' '.join(str(x) for x in (n.get('job_titles') or [])),
            ]
        ).lower()
        s = 0
        matched_keywords = 0
        if keywords:
            matched_keywords = sum(1 for kw in keywords if kw in text)
            if matched_keywords == 0:
                continue
            s += matched_keywords * 2
            if matched_keywords == len(keywords):
                s += 2
        if dep_kw and dep_kw in str(n.get('department_name', '')).lower():
            s += 2
        if cat_kw and cat_kw in str(n.get('category', '')).lower():
            s += 2
        if not keywords and not dep_kw and not cat_kw:
            s += 1
        if s == 0:
            continue
        ranked.append((s, n, matched_keywords))

    ranked.sort(key=lambda x: (x[0], x[1].get('publish_date', '')), reverse=True)
    result = []
    for s, n, matched_keywords in ranked[: max(1, min(limit, 20))]:
        result.append(
            {
                'id': n.get('id'),
                'title': n.get('title'),
                'department_name': n.get('department_name'),
                'category': n.get('category'),
                'publish_date': n.get('publish_date'),
                'content_type': n.get('content_type'),
                'score': s,
                'matched_keywords': matched_keywords,
            }
        )
    return json.dumps({'total': len(result), 'notices': result}, ensure_ascii=False, indent=2)


@mcp.tool()
async def extract_from_portal_notices(
    notice_ids: list[str],
    categories: list[str] | None = None,
    include_attachment: bool = True,
) -> str:
    """从公告站 mock 数据按公告ID定向抽取（不建议全量公告一次性抽取）。

    Args:
        notice_ids: 公告 ID 列表。
        categories: 可选。限定抽取维度。
        include_attachment: 是否拼接附件摘要信息参与抽取。
    """
    if not notice_ids:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': ['notice_ids 为空'],
            },
            ensure_ascii=False,
            indent=2,
        )
    try:
        data = _load_portal_data()
    except Exception as e:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': [str(e)],
            },
            ensure_ascii=False,
            indent=2,
        )

    notice_map = {str(n.get('id')): n for n in data.get('notices', [])}
    missing = []
    selected_texts = []
    for nid in notice_ids:
        item = notice_map.get(str(nid))
        if not item:
            missing.append(str(nid))
            continue
        selected_texts.append(_compose_notice_text(item, include_attachment))

    if not selected_texts:
        return json.dumps(
            {
                'constraints': [],
                'recommendations': [],
                'process_steps': [],
                'material_checklist': [],
                'warnings': [f'未找到有效公告: {", ".join(missing)}'],
            },
            ensure_ascii=False,
            indent=2,
        )

    merged_text = '\n\n-----\n\n'.join(selected_texts)
    result = await _extract_with_llm(merged_text, categories)
    if missing:
        result.setdefault('warnings', [])
        result['warnings'].append(f'以下公告ID未命中: {", ".join(missing)}')
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
async def list_extraction_categories() -> str:
    """返回支持的抽取维度（categories）列表，供调用方限定抽取范围。"""
    return json.dumps(
        {
            'categories': [
                'transport',
                'accommodation',
                'allowance',
                'approval',
                'audit',
                'reimbursement',
                'other',
            ],
            'stages': ['approval', 'filing', 'reimbursement'],
        },
        ensure_ascii=False,
        indent=2,
    )


if __name__ == '__main__':
    mcp.run(transport='stdio')
